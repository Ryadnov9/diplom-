import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path

# Пример расписания — ты можешь заменить это чтением из JSON-файла
json_schedule = [
    {"day": "Понеділок", "pair": 1, "group": "12-411", "subject": "Програмування мобільних", "teacher": "доц. Томалов", "type": "лекція", "link": "https://zoom.us/j/123"},
    {"day": "Понеділок", "pair": 1, "group": "12-412", "subject": "Фізика", "teacher": "проф. Іваненко", "type": "лекція", "link": ""},
    {"day": "Понеділок", "pair": 2, "group": "12-421", "subject": "Функціональний аналіз", "teacher": "доц. Кузьмич", "type": "лекція", "link": "https://zoom.us/j/123"},
    {"day": "Вівторок", "pair": 1, "group": "12-431", "subject": "Проєктування", "teacher": "доц. Литвин", "type": "практика", "link": ""},
]

# Извлекаем все группы
groups = sorted(list(set(entry["group"] for entry in json_schedule)))

# Заголовок таблицы
header = ["Дні тижня", "№ пари"] + groups

# Упорядочиваем пары по дням и группам
days_order = ["Понеділок", "Вівторок", "Середа", "Четвер", "П’ятниця"]
max_pairs = 6
table_rows = []

for day in days_order:
    for pair in range(1, max_pairs + 1):
        row = [day if pair == 1 else "", str(pair)]
        for group in groups:
            cell = ""
            for entry in json_schedule:
                if entry["day"] == day and entry["pair"] == pair and entry["group"] == group:
                    parts = [entry["subject"], entry["teacher"], entry["type"]]
                    if entry["link"]:
                        parts.append(entry["link"])
                    cell = "\n".join(filter(None, parts))
            row.append(cell)
        table_rows.append(row)

# Полная таблица
full_table = [header] + table_rows

# Создаём Excel-файл
wb = Workbook()
ws = wb.active
ws.title = "Розклад"

# Стили
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Заполняем Excel
for row_idx, row in enumerate(full_table, 1):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = alignment_center
        cell.border = border
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

# Автоматическая ширина колонок
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 4

# Сохраняем файл
wb.save("Розклад.xlsx")
print("✅ Готово: файл 'Розклад.xlsx' збережено")
