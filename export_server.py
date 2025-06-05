from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)

@app.route("/export_excel", methods=["POST"])
def export_excel():
    data = request.get_json()
    rows = data.get("table", [])

    if not rows:
        return {"error": "No table data received"}, 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Розклад"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = align
            cell.border = border

            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    # Автоширина колонок
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # Сохраняем в память и отправляем
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name="Розклад.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(debug=True)
